# # This files checks for each novel hit in the novel hits excel whether theres a km plot showing significant survival effects
import argparse
import glob
import os
from os.path import join as pjoin
import pandas as pd
from definitions import *
import numpy as np
from openpyxl import Workbook
from openpyxl.drawing.image import Image
import pickle

# def create_km_summary_excel(summary_type: str = "all"):
#     """
#     Creates an excel file summarizing the Kaplan-Meier analysis results for each cancer type and pathway.
#     For each pathway with a KM plot, it extracts the q-values for metastatic and non-metastatic patients, the pathway description, and includes the KM plot in the excel.
#     Saves the excel file in the KAPLAN_MEIER_P directory.
#     """
#     wb = Workbook()
#     ws = wb.active

#     use_cold = ['pathway', 'pathway_name', 'hr_high_vs_wildtype', 'hr_low_vs_wildtype']

#     km_cancer_directories = sorted(glob.glob(pjoin(KAPLAN_MEIER_P, '*')))

#     for km_cancer_directory in km_cancer_directories:
#         cancer_type = km_cancer_directory.split('/')[-1]

#         km_plots = sorted(glob.glob(pjoin(km_cancer_directory, '*.png')))

#         if not km_plots:
#             print(f"    No km plots were generated for {cancer_type}, skipping...")
#             continue

#         # for each km plot, add to the excel a row with cancer type, pathway id and full name, q_value_met, q_value_non_met, percentile, and km plot
#         for km_plot in km_plots:
#             pathway_id = os.path.basename(km_plot).split('_')[0]  # assuming pathway_id is the first part of the filename before the first underscore
#             pathway_name = get_pathway_description(pathway_id)
#             # get delta means from distances result csv of that cancer in the pathway_id row
#             distances_csv_path = pjoin(RESULTS_DISTANCES_P, f"{cancer_type}.csv")
#             distances_df = pd.read_csv(distances_csv_path)
#             delta_means_series = distances_df[distances_df['pathway'] == pathway_id]['delta_means']
#             delta_means = delta_means_series.values[0] if not delta_means_series.empty else np.nan

#             if summary_type == "novel":
#                 if pathway_id not in novel_pathways and pathway_id not in novel_benign_pathways:
#                     continue
#                 else:
#                     pathway_cancer_row = novel_hits_df[(novel_hits_df['pathway_id'] == pathway_id) & (novel_hits_df['cancer_type'] == cancer_type)]
#                     if pathway_cancer_row.empty:
#                         pathway_cancer_row = novel_hits_benign_df[(novel_hits_benign_df['pathway_id'] == pathway_id) & (novel_hits_benign_df['cancer_type'] == cancer_type)]
#                     if not pathway_cancer_row.empty:
#                         q_value = pathway_cancer_row['q_value'].values[0]
#                         print(f"    Adding {pathway_id} for {cancer_type} with q_value {q_value} to the summary excel")
#                         confidence_score = pathway_cancer_row['confidence_score'].values[0]
#                         relevance = pathway_cancer_row['biological_relevance'].values[0]
#                         mechanism = pathway_cancer_row['suggested_mechanism'].values[0]
#                     else:
#                         q_value = np.nan
#                         confidence_score = np.nan
#                         relevance = np.nan
#                         mechanism = np.nan
#                     ws.append([cancer_type, pathway_id, pathway_name, delta_means, f"{q_value:.3f}", f"{confidence_score:.3f}", relevance, mechanism])
#             else:
#                 ws.append([cancer_type, pathway_id, pathway_name, delta_means])

#             ws.row_dimensions[ws.max_row].height = 225  # Set height for image
#             ws.column_dimensions[km_plot_col].width = 53  # Set width for image column
#             img = Image(km_plot)
#             img.width = 400
#             img.height = 300
#             ws.add_image(img, f'{km_plot_col}{ws.max_row}')

#     # Save excel at KAPLAN_MEIER_P
#     wb.save(pjoin(KAPLAN_MEIER_P, f'km_{summary_type}_pathways_summary.xlsx'))


# def get_pathway_description(pathway_id: str) -> str:
#     with open(KEGG_PATHWAY_METADATA_FILE, 'rb') as f:
#         pathway_metadata = pickle.load(f)
#     if "cluster" in pathway_id:
#         pathway_cluster_annotations = pd.read_csv(pjoin(RESULTS_P, 'p12_pathway_cluster_annotations.csv'))
#         cluster_number = pathway_id.split('_')[-1]
#         return pathway_cluster_annotations[pathway_cluster_annotations['cluster'] == int(cluster_number)]['short_name'].values[0]
#     else:
#         return pathway_metadata.get(str(pathway_id), {}).get('name', 'Unknown')

# if __name__ == '__main__':
#     print(f"----- Summarizing KM results -----")
#     create_km_summary_excel(summary_type)
    

def create_km_summary_excel(summary_type: str = "all"):
    from openpyxl import Workbook
    from openpyxl.drawing.image import Image as XLImage
    from openpyxl.styles import Font, PatternFill, Alignment, PatternFill
    from openpyxl.utils import get_column_letter

    USE_COLS = [
        'pathway', 'pathway_name',
        'hr_high_vs_wildtype', 'hr_low_vs_wildtype', 'hr_high_vs_low',
        'q_high_vs_wildtype', 'q_low_vs_wildtype', 'p_high_vs_low',
        'n_wildtype', 'n_high', 'n_low',
    ]
    HEADER_FONT  = Font(bold=True, color='FFFFFF', name='Arial', size=10)
    BODY_FONT    = Font(name='Arial', size=9)
    COL_WIDTHS   = [15, 35, 18, 18, 18, 14, 14, 14, 12, 12, 12, 20]
    IMG_ROW_H    = 180   # points — matches roughly a thumbnail height
    IMG_COL      = len(USE_COLS) + 1   # last column index (1-based)

    def _style_header(ws, n_cols):
        for col in range(1, n_cols + 1):
            cell = ws.cell(row=1, column=col)
            cell.alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[1].height = 20

    def _write_sheet(ws, df, cancer_type, plot_dir):
        """Write dataframe rows + embedded KM plot image into worksheet ws."""
        headers = USE_COLS + ['KM Plot']
        ws.append(headers)
        _style_header(ws, len(headers))

        # Set column widths
        for i, w in enumerate(COL_WIDTHS, start=1):
            ws.column_dimensions[get_column_letter(i)].width = w
        # Image column wider
        ws.column_dimensions[get_column_letter(IMG_COL)].width = 28

        excel_row = 2
        for _, row in df.iterrows():
            row_vals = [row.get(c, '') for c in USE_COLS]
            ws.append(row_vals)

            # Style body cells
            is_sig = float(row.get('q_high_vs_wildtype', 1.0)) < 0.05
            for col in range(1, len(USE_COLS) + 1):
                cell = ws.cell(row=excel_row, column=col)
                cell.font      = BODY_FONT
                cell.alignment = Alignment(horizontal='center')

            # Embed image if it exists
            img_path = pjoin(plot_dir, f"{row['pathway']}_min_esm_log_probs_km_curve.png")
            if os.path.exists(img_path):
                ws.row_dimensions[excel_row].height = 225
                ws.column_dimensions[get_column_letter(IMG_COL)].width = 53
                img = XLImage(img_path)
                img.width = 400
                img.height = 300
                img.anchor = f"{get_column_letter(IMG_COL)}{excel_row}"
                ws.add_image(img)
            else:
                ws.cell(row=excel_row, column=IMG_COL).value = '—'

            excel_row += 1

    # ── Build workbook ────────────────────────────────────────────────────────
    wb = Workbook()
    wb.remove(wb.active)   # remove default empty sheet

    all_sig_rows = []
    km_cancer_directories = sorted(glob.glob(pjoin(KAPLAN_MEIER_P, '*')))

    for km_dir in km_cancer_directories:
        cancer_type = os.path.basename(km_dir)
        results_csv = pjoin(km_dir, 'min_esm_log_probs.csv')

        if not os.path.exists(results_csv):
            print(f"  Skipping {cancer_type}: no results CSV found.")
            continue

        df = pd.read_csv(results_csv)

        # Keep only columns that exist
        existing_cols = [c for c in USE_COLS if c in df.columns]
        df = df[existing_cols]

        # ── Per-cancer sheet (all results, image only if plot exists) ─────────
        sheet_name = cancer_type[:31]   # Excel sheet name max 31 chars
        ws = wb.create_sheet(title=sheet_name)
        _write_sheet(ws, df, cancer_type, km_dir)

        # ── Collect significant rows for the summary sheet ────────────────────
        if 'q_high_vs_wildtype' in df.columns:
            sig_df = df[df['q_high_vs_wildtype'] < 0.05].copy()
            sig_df.insert(0, 'cancer_type', cancer_type)
            all_sig_rows.append(sig_df)

    # ── Summary sheet (first sheet, significant results across all cancers) ───
    if all_sig_rows:
        summary_df = pd.concat(all_sig_rows, ignore_index=True)
        ws_summary = wb.create_sheet(title='All Significant', index=0)

        headers = ['cancer_type'] + [c for c in USE_COLS if c in summary_df.columns] + ['KM Plot']
        ws_summary.append(headers)
        _style_header(ws_summary, len(headers))

        img_col_summary = len(headers)
        ws_summary.column_dimensions[get_column_letter(img_col_summary)].width = 28
        for i, w in enumerate([18] + COL_WIDTHS, start=1):
            ws_summary.column_dimensions[get_column_letter(i)].width = w

        excel_row = 2
        for _, row in summary_df.iterrows():
            cancer = row['cancer_type']
            row_vals = [cancer] + [row.get(c, '') for c in USE_COLS if c in summary_df.columns]
            ws_summary.append(row_vals)

            for col in range(1, len(headers)):
                cell = ws_summary.cell(row=excel_row, column=col)
                cell.font      = BODY_FONT
                cell.alignment = Alignment(horizontal='center')

            # Image
            km_dir   = pjoin(KAPLAN_MEIER_P, cancer)
            img_path = pjoin(km_dir, f"{row['pathway']}_min_esm_log_probs_km_curve.png")
            if os.path.exists(img_path):
                ws_summary.row_dimensions[excel_row].height = 225
                ws_summary.column_dimensions[get_column_letter(img_col_summary)].width = 53
                img = XLImage(img_path)
                img.width = 400
                img.height = 300
                img.anchor = f"{get_column_letter(img_col_summary)}{excel_row}"
                ws_summary.add_image(img)
            else:
                ws_summary.cell(row=excel_row, column=img_col_summary).value = '—'

            excel_row += 1

    out_path = pjoin(KAPLAN_MEIER_P, f'km_{summary_type}_pathways_summary.xlsx')
    wb.save(out_path)
    print(f"Summary excel saved to: {out_path}")


if __name__ == '__main__':
    print("----- Summarizing KM results -----")
    create_km_summary_excel("all")